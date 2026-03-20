"""
Stratified sample creation with regex highlighting for human coding.
"""

import pandas as pd
import numpy as np
import re
from pathlib import Path
from typing import Dict, List, Union

try:
    from iterstrat.ml_stratifiers import MultilabelStratifiedShuffleSplit
    _HAS_ITERSTRAT = True
except ImportError:
    _HAS_ITERSTRAT = False


def count_pattern_matches(text, pattern):
    """Count regex matches in text (case-insensitive)."""
    if pd.isna(text) or not str(text).strip():
        return 0
    try:
        return len(re.findall(pattern, str(text), re.IGNORECASE))
    except Exception:
        return 0


def highlight_pattern_matches(text, pattern, max_length=200):
    """Return up to 3 matched snippets with context for visual inspection."""
    if pd.isna(text) or not str(text).strip():
        return ''
    try:
        matches = list(re.finditer(pattern, str(text), re.IGNORECASE))
        if not matches:
            return ''
        snippets = []
        for match in matches[:3]:
            start = max(0, match.start() - 20)
            end = min(len(text), match.end() + 20)
            snippets.append(f"...{text[start:end].strip()}...")
        result = ' | '.join(snippets)
        return result[:max_length] if len(result) > max_length else result
    except Exception:
        return ''


def create_stratified_sample(df: pd.DataFrame, patterns: Dict[str, str],
                             text_cols: List[str] = None,
                             coding_labels: List[str] = None,
                             sample_size: float = 0.2,
                             random_seed: int = 42,
                             semantic_threshold: Union[str, float] = 'median') -> pd.DataFrame:
    """
    Create a stratified random sample using iterative stratification.

    Stratifies on binary regex flags and semantic score levels per criterion
    using Sechidis et al. (2011) iterative stratification via iterstrat.

    Args:
        df: Input DataFrame with binary criterion columns and optional
            ``{criterion}_semantic_max`` columns.
        patterns: Dict of {criterion: regex_pattern} for highlighting.
        text_cols: Columns to combine for pattern matching
            (default: title, abstract, keywords).
        coding_labels: Label columns to add for human coding
            (default: pattern keys).
        sample_size: Proportion to sample (default 0.2).
        random_seed: For reproducibility.
        semantic_threshold: Threshold for binarizing semantic scores.
            'median' (default) uses per-criterion median of non-zero scores.
            A float (e.g., 0.5) applies a fixed threshold across all criteria.

    Returns:
        DataFrame with pattern highlights and empty coding columns.
    """
    if not _HAS_ITERSTRAT:
        raise ImportError(
            "iterstrat is required for stratified sampling. "
            "Install: pip install iterative-stratification"
        )

    if text_cols is None:
        text_cols = ['title', 'abstract', 'keywords']
    if coding_labels is None:
        coding_labels = list(patterns.keys())

    available_text_cols = [c for c in text_cols if c in df.columns]
    df = df.copy()
    df['_combined_text'] = df[available_text_cols].fillna('').astype(str).agg(' '.join, axis=1)

    # Add pattern count and snippet columns
    for label, pattern in patterns.items():
        df[f'{label}_pattern_count'] = df['_combined_text'].apply(
            lambda x, p=pattern: count_pattern_matches(x, p))
        df[f'{label}_pattern_snippets'] = df['_combined_text'].apply(
            lambda x, p=pattern: highlight_pattern_matches(x, p))

    # Build binary label matrix for iterative stratification
    criteria = list(patterns.keys())
    strat_cols = []

    # Binary regex flags
    for c in criteria:
        col = f'_strat_{c}'
        df[col] = (df[c].fillna(0).astype(int) if c in df.columns
                   else (df[f'{c}_pattern_count'] > 0).astype(int))
        strat_cols.append(col)

    # Binary semantic score levels (high/low)
    for c in criteria:
        sem_col = f'{c}_semantic_max'
        strat_col = f'_strat_{c}_sem_high'
        if sem_col in df.columns:
            scores = df[sem_col].fillna(0)
            nonzero = scores[scores > 0]
            if len(nonzero) > 0:
                if semantic_threshold == 'median':
                    thresh = nonzero.median()
                else:
                    thresh = float(semantic_threshold)
                df[strat_col] = (scores >= thresh).astype(int)
                strat_cols.append(strat_col)

    # Iterative stratification
    y = df[strat_cols].values
    X = np.arange(len(df)).reshape(-1, 1)
    msss = MultilabelStratifiedShuffleSplit(
        n_splits=1, test_size=1 - sample_size, random_state=random_seed
    )
    sample_idx, _ = next(msss.split(X, y))
    sample_df = df.iloc[sample_idx].reset_index(drop=True)

    print(f"Created sample: {len(sample_df)} records ({sample_size * 100:.0f}%)")
    for col in strat_cols:
        label = col.replace('_strat_', '')
        pos = sample_df[col].sum()
        pct = pos / len(sample_df) * 100
        pop_pct = df[col].sum() / len(df) * 100
        print(f"  {label:30s}: {pos:4d} ({pct:5.1f}%) [population: {pop_pct:.1f}%]")

    # Add empty coding columns
    for label in coding_labels:
        sample_df[label] = ''
    sample_df['notes'] = ''
    sample_df['coder_id'] = ''
    sample_df['coding_date'] = ''

    # Clean up temp columns
    temp_cols = ['_combined_text'] + [c for c in sample_df.columns if c.startswith('_strat_')]
    sample_df = sample_df.drop(columns=temp_cols, errors='ignore')

    return sample_df


def apply_conditional_formatting(excel_file, patterns: Dict[str, str]):
    """
    Apply conditional formatting to Excel coding sheet.

    Highlights coding column headers green and pattern count cells > 0 yellow.
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl required for conditional formatting. Install: pip install openpyxl")
        return

    wb = load_workbook(excel_file)
    ws = wb.active
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    bold_font = Font(bold=True)

    header_row = {cell.value: cell.column for cell in ws[1]}

    # Green headers for coding columns
    for label in patterns:
        if label in header_row:
            col_letter = get_column_letter(header_row[label])
            ws[f'{col_letter}1'].fill = green_fill
            ws[f'{col_letter}1'].font = bold_font

    # Yellow highlight for count > 0
    for label in patterns:
        count_col = f'{label}_pattern_count'
        if count_col in header_row:
            col_letter = get_column_letter(header_row[count_col])
            for row in range(2, ws.max_row + 1):
                cell = ws[f'{col_letter}{row}']
                try:
                    if cell.value and int(cell.value) > 0:
                        cell.fill = yellow_fill
                        cell.font = bold_font
                except (ValueError, TypeError):
                    pass

    # Column widths
    width_hints = {'title': 50, 'abstract': 80, 'keywords': 40, 'notes': 40}
    for label in patterns:
        width_hints[f'{label}_pattern_snippets'] = 60

    for col_name, width in width_hints.items():
        if col_name in header_row:
            ws.column_dimensions[get_column_letter(header_row[col_name])].width = width

    ws.freeze_panes = 'D2'
    wb.save(excel_file)
    print(f"Conditional formatting applied: {excel_file}")


def save_sample_excel(sample_df: pd.DataFrame, output_file, patterns: Dict[str, str]):
    """Save sample to Excel with conditional formatting."""
    output_file = Path(output_file)
    sample_df.to_excel(output_file, index=False, sheet_name='Sample')
    apply_conditional_formatting(output_file, patterns)
    print(f"Sample saved: {output_file}")
